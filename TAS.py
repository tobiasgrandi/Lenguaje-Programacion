import openpyxl

RUTA_EXCEL = "TAS.xlsx"

EXCEL = openpyxl.load_workbook(RUTA_EXCEL).active


class TAS():

    def __init__(self):
        self.primera_columna = obtener_primera_columna(EXCEL)
        self.primera_fila = obtener_primera_fila(EXCEL)

    def obtener_produccion(self, valor_columna, valor_fila):
        columna = obtener_posicion_columna(self.primera_fila, valor_fila)
        fila = obtener_posicion_fila(self.primera_columna, valor_columna)

        produccion = EXCEL.cell(row=fila, column=columna).value

        return produccion


def obtener_posicion_columna(fila, valor_fila):

    index_valor = fila.index(valor_fila)
    return index_valor + 1

def obtener_posicion_fila(columna, valor_columna):

    index_valor = columna.index(valor_columna)
    return index_valor + 1

def obtener_primera_columna(hoja_excel):

    columna = []

    for fila in hoja_excel.iter_rows(min_row = 1, max_row = hoja_excel.max_row, min_col = 1, max_col = 1):
        celda = fila[0]
        columna.append(celda.value)

    return columna

def obtener_primera_fila(hoja_excel):

    fila = []

    for columnas in hoja_excel.iter_rows(min_row=1, max_row= 1, min_col=1, max_col= hoja_excel.max_column):
        for celda in columnas:
            fila.append(celda.value)

    return fila